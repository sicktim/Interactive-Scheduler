"""
Extract structured event data from the CF PHASE sheet of Continuity_25A.xlsx.
Outputs continuity-CF.json with all events organized by section.
"""

import json
import re
from openpyxl import load_workbook

WORKBOOK = 'C:/Users/sickt/OneDrive/Documents/Claude/tps_schedule_gannt/Data-Extract/Continuity/Continuity_25A.xlsx'
OUTPUT = 'C:/Users/sickt/OneDrive/Documents/Claude/tps_schedule_gannt/Data-Extract/Continuity/continuity-CF.json'

def parse_duration(val):
    """Convert duration value to decimal hours.
    Formats: 1.5 (float), '1+05' (H+MM), '(2.0*)' (float with note), 1 (int)
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return round(float(val), 3)
    s = str(val).strip()
    # Remove parentheses and asterisks: "(2.0*)" -> "2.0"
    s = re.sub(r'[()* ]', '', s)
    # H+MM format
    m = re.match(r'^(\d+)\+(\d{2})$', s)
    if m:
        hours = int(m.group(1))
        minutes = int(m.group(2))
        return round(hours + minutes / 60.0, 3)
    # Plain number
    try:
        return round(float(s), 3)
    except ValueError:
        return None


def is_section_header(row_data):
    """A section header has text ONLY in column B (no code in A, no crew in C, minimal other cols).
    Known section headers from the data."""
    a, b, c, d, e, f, g, h, i, j, k = row_data
    if a is not None:
        return False
    if b is None:
        return False
    # Must have no crew, no config, no time, no LOX, no quals, no capes
    if c is not None or d is not None or e is not None:
        return False
    if h is not None or i is not None or j is not None or k is not None:
        return False
    # Can have notes/prereqs if it's a continuation row -- but section headers typically don't
    # Check known patterns
    b_stripped = str(b).strip()
    known_sections = [
        'TPS AIRMANSHIP', 'Glider', 'T-38', 'F-16', 'C-12', 'RPA',
        'Learjet', 'Photo / Safety Chase', 'Photo/Safety Chase'
    ]
    if b_stripped in known_sections:
        return True
    # Also check if it has notes -- if so, it's a continuation, not a section header
    if f is not None or g is not None:
        return False
    return False


def is_event_start(row_data):
    """An event starts with a code in column A."""
    a = row_data[0]
    if a is None:
        return False
    s = str(a).strip()
    # Event codes start with "CF " followed by digits
    return s.startswith('CF ')


def is_empty_row(row_data):
    """All cells are None."""
    return all(v is None for v in row_data)


def is_variant_row(row_data):
    """A continuation row that has an aircraft variant in column B (e.g., 'T-38C', 'F-16D', 'LJ-25', 'C-12C')."""
    a, b = row_data[0], row_data[1]
    if a is not None:
        return False
    if b is None:
        return False
    b_str = str(b).strip()
    variant_patterns = [
        'T-38C', 'F-16D', 'LJ-25', 'C-12C', 'C-172',
        'SGS 2-33', 'ASK-21', 'G-103', 'UTD',
        'F-16D & UTD'
    ]
    for pat in variant_patterns:
        if b_str.startswith(pat):
            return True
    return False


def strip_val(v):
    """Strip string values, return None for empty."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def extract_cf_phase():
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb['CF PHASE']

    # Read all rows (skip header row 1)
    all_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=11, values_only=False):
        row_num = row[0].row
        vals = tuple(cell.value for cell in row)
        all_rows.append((row_num, vals))

    sections = []
    current_section = None
    current_event = None
    events_in_section = []

    def finalize_event():
        """Flush the current event into the section's event list."""
        nonlocal current_event
        if current_event is not None:
            events_in_section.append(current_event)
            current_event = None

    def finalize_section():
        """Flush the current section."""
        nonlocal current_section, events_in_section
        finalize_event()
        if current_section is not None and events_in_section:
            sections.append({
                "name": current_section,
                "events": events_in_section
            })
        elif current_section is not None:
            # Section with no events (shouldn't happen, but handle)
            sections.append({
                "name": current_section,
                "events": []
            })
        events_in_section = []

    for row_num, vals in all_rows:
        a, b, c, d, e, f, g, h, i, j, k = vals

        # Skip empty rows
        if is_empty_row(vals):
            continue

        # Check for section header
        if is_section_header(vals):
            finalize_section()
            current_section = str(b).strip()
            continue

        # Check for event start (code in column A)
        if is_event_start(vals):
            finalize_event()
            event_code = str(a).strip()
            mission_name = strip_val(b)

            # Determine aircraft type from LOX column, fall back to section name
            aircraft_type = strip_val(h)
            if aircraft_type is None and current_section:
                # Infer from section name for sections that map directly to aircraft
                section_aircraft_map = {
                    'Learjet': 'Learjet',
                    'TPS AIRMANSHIP': None,  # C-172 but already in LOX
                    'Glider': None,  # already in LOX
                }
                aircraft_type = section_aircraft_map.get(current_section)

            # Config
            config = strip_val(d)

            # Duration
            duration = parse_duration(e)

            # Crew
            crew = []
            c_val = strip_val(c)
            if c_val:
                crew.append(str(c_val))

            # Notes
            notes = []
            f_val = strip_val(f)
            if f_val:
                notes.append(str(f_val))

            # Prerequisites
            prereqs = []
            g_val = strip_val(g)
            if g_val:
                prereqs.append(str(g_val))

            # Quals and capes
            aircrew_quals = strip_val(i)
            req_cape = strip_val(j)
            des_cape = strip_val(k)

            current_event = {
                "eventCode": event_code,
                "missionName": mission_name,
                "section": current_section,
                "aircraftType": aircraft_type,
                "aircraftVariant": None,
                "config": config,
                "duration": duration,
                "crew": crew,
                "notes": notes,
                "prerequisites": prereqs,
                "aircrewQuals": aircrew_quals,
                "aircraftRequiredCape": req_cape,
                "aircraftDesiredCape": des_cape,
                "sourceRows": [row_num]
            }
            continue

        # Continuation row (no code in A, not a section header, not empty)
        if current_event is not None:
            current_event["sourceRows"].append(row_num)

            # Check for aircraft variant in B column
            b_val = strip_val(b)
            if b_val:
                # Known variant patterns
                variant_strs = ['T-38C', 'F-16D', 'LJ-25', 'C-12C', 'C-172',
                                'SGS 2-33', 'ASK-21', 'G-103', 'UTD',
                                'F-16D & UTD', 'C-12C v C-12C']
                matched_variant = False
                for vs in variant_strs:
                    if b_val.startswith(vs):
                        # Set variant (keep first one if multiple continuation rows)
                        if current_event["aircraftVariant"] is None:
                            current_event["aircraftVariant"] = b_val
                        matched_variant = True
                        break
                if not matched_variant:
                    # Could be a sub-label like "Photo / Safety Chase" appearing as continuation
                    # or additional mission info -- store as note if meaningful
                    # But check if it's the "ASK-21/G-103* (Hvy Studs)" pattern
                    if 'ASK-21' in b_val or 'G-103' in b_val:
                        if current_event["aircraftVariant"] is None:
                            current_event["aircraftVariant"] = b_val
                    # Otherwise skip non-variant B column text in continuations
                    # (it may repeat section name or be irrelevant)

            # Crew from C column
            c_val = strip_val(c)
            if c_val:
                current_event["crew"].append(str(c_val))

            # Config from D column (merge if present)
            d_val = strip_val(d)
            if d_val:
                if current_event["config"] is None:
                    current_event["config"] = d_val
                elif d_val not in current_event["config"]:
                    current_event["config"] += " / " + d_val

            # Duration from E column -- pick up if primary didn't have one,
            # or note alternate durations
            e_val = parse_duration(e)
            if e_val is not None:
                if current_event["duration"] is None:
                    current_event["duration"] = e_val
                # If continuation has a parenthetical duration like "(2.0*)", it's an alternate
                # We'll note it but keep the primary duration
                elif isinstance(e, str) and '(' in str(e):
                    current_event["notes"].append(f"Alternate duration: {str(e).strip()}")

            # Notes from F column
            f_val = strip_val(f)
            if f_val:
                current_event["notes"].append(str(f_val))

            # Prerequisites from G column
            g_val = strip_val(g)
            if g_val:
                current_event["prerequisites"].append(str(g_val))

            # Quals
            i_val = strip_val(i)
            if i_val and current_event["aircrewQuals"] is None:
                current_event["aircrewQuals"] = i_val

            # Required cape
            j_val = strip_val(j)
            if j_val and current_event["aircraftRequiredCape"] is None:
                current_event["aircraftRequiredCape"] = j_val

            # Desired cape
            k_val = strip_val(k)
            if k_val and current_event["aircraftDesiredCape"] is None:
                current_event["aircraftDesiredCape"] = k_val

    # Flush final event and section
    finalize_section()

    # Collect aircraft types across all events
    aircraft_types = set()
    total_events = 0
    for sec in sections:
        total_events += len(sec["events"])
        for ev in sec["events"]:
            if ev["aircraftType"]:
                aircraft_types.add(ev["aircraftType"])

    output = {
        "sheet": "CF PHASE",
        "classCode": "25A",
        "extractedAt": "2026-02-19",
        "sections": sections,
        "summary": {
            "totalEvents": total_events,
            "totalSections": len(sections),
            "aircraftTypes": sorted(list(aircraft_types))
        }
    }

    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"Wrote {OUTPUT}")
    print(f"Total sections: {len(sections)}")
    print(f"Total events: {total_events}")
    print(f"Aircraft types: {sorted(aircraft_types)}")
    for sec in sections:
        print(f"  Section '{sec['name']}': {len(sec['events'])} events")
        for ev in sec['events']:
            print(f"    {ev['eventCode']} - {ev['missionName']} (rows {ev['sourceRows']})")


if __name__ == '__main__':
    extract_cf_phase()
